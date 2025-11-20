"""Utility functions for extracting text from uploaded documents."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from pathlib import Path
from typing import Iterable, List

try:
    import fitz  # type: ignore
except ImportError:  # pragma: no cover - optional dependency guard
    fitz = None

from docx import Document as DocxDocument
from openpyxl import load_workbook
from PyPDF2 import PdfReader


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff"}


def normalise_filename(name: str) -> str:
    """Remove directory components and non filename-safe chars."""

    base = Path(name).name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return cleaned or "file"


def extract_text(path: Path) -> str:
    """Extract best-effort plain text from the file."""

    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".docx":
        return _extract_docx(path)
    if suffix == ".xlsx":
        return _extract_xlsx(path)
    if suffix == ".csv":
        return _extract_csv(path)
    if suffix == ".pptx":
        return _extract_pptx(path)
    return ""


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 100) -> List[str]:
    """Split text into overlapping character chunks."""

    if not text:
        return []

    cleaned = text.replace("\r\n", "\n").strip()
    if not cleaned:
        return []

    chunks: List[str] = []
    step = max(1, chunk_size - overlap)
    for start in range(0, len(cleaned), step):
        chunk = cleaned[start : start + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
    return chunks


def _extract_pdf(path: Path) -> str:
    pages: List[str] = []

    if fitz is not None:
        try:
            with fitz.open(path) as doc:
                for idx, page in enumerate(doc, start=1):
                    text = page.get_text("text") or ""
                    if text.strip():
                        pages.append(f"--- Page {idx} ---\n{text.strip()}")
        except Exception:
            pages = []

    if not pages:
        reader = PdfReader(str(path))
        for idx, page in enumerate(reader.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            if text.strip():
                pages.append(f"--- Page {idx} ---\n{text.strip()}")

    return "\n\n".join(pages).strip()


def _extract_docx(path: Path) -> str:
    doc = DocxDocument(str(path))
    parts = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n".join(parts).strip()


def _extract_xlsx(path: Path) -> str:
    try:
        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception:
        return ""

    lines: List[str] = []
    try:
        for sheet in workbook.worksheets[:10]:
            lines.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True):
                values = ["" if cell is None else str(cell) for cell in row]
                if any(values):
                    lines.append("\t".join(values))
    finally:
        workbook.close()
    return "\n".join(lines).strip()


def _extract_csv(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        return "\n".join(",".join(row) for row in reader).strip()


def _extract_pptx(path: Path) -> str:
    lines: List[str] = []
    with zipfile.ZipFile(path, "r") as pptx:
        for filename in pptx.namelist():
            if not filename.startswith("ppt/slides/slide") or not filename.endswith(".xml"):
                continue
            with pptx.open(filename) as slide_file:
                data = slide_file.read().decode("utf-8", errors="ignore")
                text = _strip_xml_tags(data)
                if text:
                    lines.append(text)
    return "\n\n".join(lines).strip()


def _strip_xml_tags(xml_text: str) -> str:
    return re.sub(r"<[^>]+>", " ", xml_text).replace("&amp;", "&").strip()

